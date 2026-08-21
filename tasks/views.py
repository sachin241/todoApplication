from datetime import date, datetime
from io import BytesIO

from django.contrib import messages
from django.db.models import Case, IntegerField, Value, When
from django.http import HttpResponse, JsonResponse
from django.shortcuts import get_object_or_404, redirect, render
from django.urls import reverse
from django.utils import timezone
from django.views.decorators.http import require_POST
from openpyxl import Workbook, load_workbook

from .forms import ImportForm, TaskForm
from .models import Task


def filtered_tasks(request):
    tasks = Task.objects.all()
    status = request.GET.get("status", "")
    priority = request.GET.get("priority", "")
    search = request.GET.get("q", "").strip()
    sort = request.GET.get("sort", "due_date")
    if status in Task.Status.values:
        tasks = tasks.filter(status=status)
    if priority in Task.Priority.values:
        tasks = tasks.filter(priority=priority)
    if search:
        tasks = tasks.filter(title__icontains=search)
    ordering = {
        "due_date": ("due_date", "-created_at"),
        "created_at": ("-created_at",),
        "priority": (),
    }
    if sort == "priority":
        tasks = tasks.annotate(priority_rank=Case(
            When(priority=Task.Priority.HIGH, then=Value(1)),
            When(priority=Task.Priority.MEDIUM, then=Value(2)),
            default=Value(3), output_field=IntegerField(),
        )).order_by("priority_rank", "due_date", "-created_at")
    else:
        sort = sort if sort in ordering else "due_date"
        tasks = tasks.order_by(*ordering[sort])
    return tasks, {"status": status, "priority": priority, "q": search, "sort": sort}


def task_list(request):
    tasks, filters = filtered_tasks(request)
    today = timezone.localdate()
    priority_pop = Task.objects.filter(priority=Task.Priority.HIGH, due_date=today).exclude(status=Task.Status.COMPLETED)
    overdue = Task.objects.filter(due_date__lt=today).exclude(status=Task.Status.COMPLETED).order_by("due_date")
    return render(request, "tasks/task_list.html", {
        **filters, "tasks": tasks, "priority_pop": priority_pop, "overdue": overdue,
        "status_choices": Task.Status.choices, "priority_choices": Task.Priority.choices,
    })


def task_create(request):
    form = TaskForm(request.POST or None, initial={"due_date": request.GET.get("due_date") or None})
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Task created successfully.")
        return redirect("tasks:list")
    return render(request, "tasks/task_form.html", {"form": form, "page_title": "Create task"})


def task_edit(request, pk):
    task = get_object_or_404(Task, pk=pk)
    form = TaskForm(request.POST or None, instance=task)
    if request.method == "POST" and form.is_valid():
        form.save()
        messages.success(request, "Task updated successfully.")
        return redirect("tasks:list")
    return render(request, "tasks/task_form.html", {"form": form, "task": task, "page_title": "Edit task"})


def task_delete(request, pk):
    task = get_object_or_404(Task, pk=pk)
    if request.method == "POST":
        task.delete()
        messages.success(request, "Task deleted.")
        return redirect("tasks:list")
    return render(request, "tasks/task_confirm_delete.html", {"task": task})


@require_POST
def task_toggle(request, pk):
    task = get_object_or_404(Task, pk=pk)
    task.status = Task.Status.PENDING if task.status == Task.Status.COMPLETED else Task.Status.COMPLETED
    task.save(update_fields=["status", "updated_at"])
    messages.success(request, f'“{task.title}” marked {task.get_status_display().lower()}.')
    return redirect(request.POST.get("next") or "tasks:list")


def calendar_view(request):
    return render(request, "tasks/calendar.html")


def calendar_events(request):
    colors = {Task.Priority.HIGH: "#dc3545", Task.Priority.MEDIUM: "#ffc107", Task.Priority.LOW: "#198754"}
    events = []
    for task in Task.objects.exclude(due_date__isnull=True):
        events.append({
            "id": task.pk, "title": task.title, "start": task.due_date.isoformat(), "allDay": True,
            "color": colors[task.priority], "url": reverse("tasks:edit", args=[task.pk]),
            "extendedProps": {"description": task.description, "status": task.get_status_display(), "priority": task.get_priority_display(), "editUrl": reverse("tasks:edit", args=[task.pk]), "toggleUrl": reverse("tasks:toggle", args=[task.pk])},
        })
    return JsonResponse(events, safe=False)


def export_tasks(request):
    tasks, _ = filtered_tasks(request)
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Tasks"
    sheet.append(["Title", "Description", "Status", "Priority", "Due Date", "Created At"])
    for task in tasks:
        sheet.append([task.title, task.description, task.get_status_display(), task.get_priority_display(), task.due_date, task.created_at.replace(tzinfo=None)])
    for column in sheet.columns:
        sheet.column_dimensions[column[0].column_letter].width = min(max(len(str(cell.value or "")) for cell in column) + 2, 45)
    output = BytesIO()
    workbook.save(output)
    response = HttpResponse(output.getvalue(), content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response["Content-Disposition"] = 'attachment; filename="tasks_export.xlsx"'
    return response


def parse_date(value):
    if value in (None, ""):
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if isinstance(value, str):
        for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%m/%d/%Y"):
            try:
                return datetime.strptime(value.strip(), fmt).date()
            except ValueError:
                pass
    raise ValueError("Due Date must be a valid date.")


def import_tasks(request):
    form = ImportForm(request.POST or None, request.FILES or None)
    if request.method == "POST" and form.is_valid():
        try:
            workbook = load_workbook(form.cleaned_data["file"], read_only=True, data_only=True)
            sheet = workbook.active
        except Exception:
            form.add_error("file", "The uploaded file could not be read as an Excel workbook.")
        else:
            headers = [str(cell.value or "").strip().lower() for cell in next(sheet.iter_rows(min_row=1, max_row=1))]
            required = ["title", "description", "status", "priority", "due date", "created at"]
            if headers[:6] != required:
                form.add_error("file", "Header row must be: Title, Description, Status, Priority, Due Date, Created At.")
            else:
                status_map = {label.lower(): value for value, label in Task.Status.choices}
                priority_map = {label.lower(): value for value, label in Task.Priority.choices}
                pending, skipped = [], []
                for row_number, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    if not any(value not in (None, "") for value in row):
                        continue
                    try:
                        title = str(row[0] or "").strip()
                        if not title:
                            raise ValueError("Title is required.")
                        if len(title) > 255:
                            raise ValueError("Title must be 255 characters or fewer.")
                        status = status_map.get(str(row[2] or "").strip().lower())
                        priority = priority_map.get(str(row[3] or "").strip().lower())
                        if not status:
                            raise ValueError("Status must be Pending, In Progress, or Completed.")
                        if not priority:
                            raise ValueError("Priority must be Low, Medium, or High.")
                        pending.append(Task(title=title, description=str(row[1] or "").strip(), status=status, priority=priority, due_date=parse_date(row[4])))
                    except (ValueError, IndexError) as exc:
                        skipped.append(f"Row {row_number}: {exc}")
                Task.objects.bulk_create(pending)
                messages.success(request, f"Imported {len(pending)} task(s); skipped {len(skipped)} row(s).")
                if skipped:
                    messages.warning(request, " ".join(skipped))
                return redirect("tasks:list")
    return render(request, "tasks/import_tasks.html", {"form": form})
